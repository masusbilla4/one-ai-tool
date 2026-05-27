# -*- coding: utf-8 -*-
"""ASR Aligner - Flask routes (based on actual ASR Web App)."""
import os
import io
import json
import re
import math
import time
import random
import datetime
import threading
from flask import Blueprint, render_template, request, jsonify, session, send_file, flash, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font as OpenPyXLFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock

from .alignment_engine import run_alignment, align_translation_local
from auth.routes import login_required
from settings.routes import get_gemini_api_key

asr_bp = Blueprint('asr', __name__, template_folder='templates')

# Get Gemini API key from session or config
def get_api_key():
    """Get Gemini API key from session or fallback to config."""
    return get_gemini_api_key()

# Configuration (from actual ASR web app)
DEFAULT_MODEL = "gemini-2.5-flash"
AVAILABLE_MODELS = [
    "gemini-2.5-flash", "gemini-2.5-flash-lite", "gemini-2.5-pro",
    "gemini-2.0-flash", "gemini-1.5-flash", "gemini-1.5-pro"
]
MODEL_RPM_LIMITS = {
    "gemini-2.5-flash": 15, "gemini-2.5-flash-lite": 15,
    "gemini-2.5-pro": 10, "gemini-2.0-flash": 15,
    "gemini-1.5-flash": 15, "gemini-1.5-pro": 10
}
BATCH_SIZE = 40
OVERLAP_CONTEXT = 3
RETRY_DELAY = 10
MAX_RETRIES = 5
BATCH_DELAY = 15

# Server-side stored alignment data
_stored_alignment = {"data": None, "overall_wer": 0, "overall_stats": {}}

# Async alignment store
_align_store = {"result": None, "error": None, "done": False, "task_id": 0}

# Re-evaluate store
_reeval_store = {"result": None, "error": None, "done": False, "task_id": 0}

# Progress store for AI tasks
_progress_store = {
    "logs": [], "progress": 0, "status": "", "running": False,
    "task_id": 0, "partial_data": None, "partial_count": 0
}

def add_log(msg):
    _progress_store["logs"].append(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")
    if len(_progress_store["logs"]) > 500:
        _progress_store["logs"] = _progress_store["logs"][-500:]

def reset_progress(task_name=""):
    _progress_store["task_id"] += 1
    _progress_store["logs"] = []
    _progress_store["progress"] = 0
    _progress_store["status"] = task_name
    _progress_store["running"] = True
    if task_name:
        add_log(f"▶ Started: {task_name}")

def finish_progress(msg="Done"):
    _progress_store["running"] = False
    _progress_store["progress"] = 100
    add_log(f"✅ {msg}")


@asr_bp.route('/')
@login_required
def asr_home():
    """ASR Aligner home page."""
    return render_template('asr/aligner.html')


@asr_bp.route('/upload-file', methods=['POST'])
@login_required
def upload_file():
    """Upload and extract text from file for ASR alignment."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    lines = []
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.txt'):
            content = file.read().decode('utf-8')
            lines = [l.strip() for l in content.split('\n') if l.strip()]
        elif filename.endswith('.srt'):
            content = file.read().decode('utf-8')
            # Extract subtitle text from SRT
            lines = []
            for line in content.split('\n'):
                line = line.strip()
                if line and not line.isdigit() and '-->' not in line:
                    lines.append(line)
        elif filename.endswith('.csv'):
            import csv
            from io import TextIOWrapper
            content = file.read().decode('utf-8')
            reader = csv.reader(content.splitlines())
            for row in reader:
                if row:
                    lines.append(row[0].strip())
        elif filename.endswith('.xlsx'):
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(filename=BytesIO(file.read()))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0]:
                    lines.append(str(row[0]).strip())
        else:
            return jsonify({"error": "Unsupported file format"}), 400
        
        return jsonify({"lines": lines, "count": len(lines)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@asr_bp.route('/align', methods=['POST'])
@login_required
def api_align():
    """Run ASR alignment (async)."""
    data = request.json
    true_text = data.get('true_text', '').strip()
    asr_text = data.get('asr_text', '').strip()

    if not true_text or not asr_text:
        return jsonify({"error": "Both True Text and ASR Result are required."}), 400

    true_lines = [l.strip() for l in true_text.split("\n") if l.strip()]
    asr_lines = [l.rstrip() for l in asr_text.splitlines() if l.strip()]

    # Word count mismatch warning
    ref_wc = sum(len(l.split()) for l in true_lines)
    hyp_wc = sum(len(l.split()) for l in asr_lines)
    if ref_wc > 0 and hyp_wc > 0:
        ratio = max(ref_wc, hyp_wc) / min(ref_wc, hyp_wc)
        if ratio > 3:
            return jsonify({
                "warning": True,
                "message": f"Significant word count mismatch: True Text={ref_wc}, ASR={hyp_wc}, Ratio={ratio:.1f}x"
            })

    # Run alignment in background thread
    _align_store["result"] = None
    _align_store["error"] = None
    _align_store["done"] = False
    _align_store["task_id"] += 1
    task_id = _align_store["task_id"]

    def _run():
        try:
            result = run_alignment(true_lines, asr_lines)
            _align_store["result"] = result
            _align_store["done"] = True
        except Exception as e:
            _align_store["error"] = str(e)
            _align_store["done"] = True

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"status": "started", "task_id": task_id})


@asr_bp.route('/align-result')
@login_required
def api_align_result():
    """Poll for async alignment result."""
    if _align_store["done"]:
        if _align_store["error"]:
            return jsonify({"done": True, "error": _align_store["error"]})
        result = _align_store["result"]
        _stored_alignment["data"] = result.get("alignment_data", [])
        _stored_alignment["overall_wer"] = result.get("overall_wer", 0)
        _stored_alignment["overall_stats"] = result.get("overall_stats", {})
        return jsonify({"done": True, "result": result})
    return jsonify({"done": False})


@asr_bp.route('/align/translation', methods=['POST'])
@login_required
def api_align_translation_local():
    """Align translation text to ASR segments (local method)."""
    data = request.json
    trans_blob = data.get("translation_text", "").strip()
    alignment_data = data.get("alignment_data", [])

    if not trans_blob:
        return jsonify({"error": "Translation text is required."}), 400
    if not alignment_data:
        return jsonify({"error": "Run ASR alignment first."}), 400

    translations, method = align_translation_local(alignment_data, trans_blob)

    for i, trans in enumerate(translations):
        if i < len(alignment_data):
            alignment_data[i]["translation"] = trans

    return jsonify({
        "alignment_data": alignment_data,
        "method": method,
        "count": len(translations)
    })


@asr_bp.route('/reevaluate', methods=['POST'])
@login_required
def api_reevaluate():
    """Re-run local diff/WER calculation."""
    alignment_data = _stored_alignment.get("data")
    if not alignment_data:
        return jsonify({"error": "No stored alignment data. Run alignment first."}), 400

    _reeval_store["result"] = None
    _reeval_store["error"] = None
    _reeval_store["done"] = False
    _reeval_store["task_id"] += 1
    task_id = _reeval_store["task_id"]

    def _run():
        try:
            total_subs = 0
            total_dels = 0
            total_ins = 0
            total_refs = 0

            for d in alignment_data:
                true_text = d.get("true", "")
                asr_text = d.get("asr_nobreak", "") or d.get("asr", "")
                true_words = true_text.split()
                asr_words = asr_text.split()
                d["count"] = len(true_words)

                result = run_alignment([true_text], [asr_text])
                if result["alignment_data"]:
                    rd = result["alignment_data"][0]
                    d["diffs"] = rd.get("diffs", [])
                    d["wrong_count"] = rd.get("wrong_count", 0)
                    d["wrong_list"] = rd.get("wrong_list", "")
                    d["wer"] = rd.get("wer", 0)
                    d["srr"] = rd.get("srr", "")
                    d["score"] = rd.get("score", "")

                total_refs += len(true_words)
                stats = result.get("overall_stats", {})
                total_subs += stats.get("subs", 0)
                total_dels += stats.get("dels", 0)
                total_ins += stats.get("ins", 0)

            overall_wer = round((total_subs + total_dels + total_ins) / total_refs * 100, 1) if total_refs else 0
            overall_stats = {"subs": total_subs, "dels": total_dels, "ins": total_ins, "refs": total_refs}

            _reeval_store["result"] = {
                "alignment_data": alignment_data,
                "overall_wer": overall_wer,
                "overall_stats": overall_stats
            }
            _reeval_store["done"] = True
        except Exception as e:
            _reeval_store["error"] = str(e)
            _reeval_store["done"] = True

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"status": "started", "task_id": task_id})


@asr_bp.route('/reevaluate-result')
@login_required
def api_reevaluate_result():
    """Poll for async reevaluate result."""
    if _reeval_store["done"]:
        if _reeval_store["error"]:
            return jsonify({"done": True, "error": _reeval_store["error"]})
        return jsonify({"done": True, "result": _reeval_store["result"]})
    return jsonify({"done": False})


@asr_bp.route('/export/csv/<task_id>')
@login_required
def export_csv(task_id):
    """Export alignment results as CSV."""
    import csv
    from io import StringIO, BytesIO
    
    alignment_data = _stored_alignment.get("data")
    if not alignment_data:
        flash('No data to export.', 'error')
        return redirect(url_for('asr.asr_home'))

    csv_buffer = StringIO()
    fieldnames = ['id', 'true', 'count', 'asr_separated', 'wer', 'score', 'translation', 'ai_reason']
    writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()

    for row in alignment_data:
        writer.writerow({
            'id': row['id'],
            'true': row['true'],
            'count': row['count'],
            'asr_separated': row['asr_separated'],
            'wer': row['wer'],
            'score': row['score'],
            'translation': row.get('translation', ''),
            'ai_reason': row.get('ai_reason', '')
        })

    output = BytesIO()
    output.write(csv_buffer.getvalue().encode('utf-8'))
    output.seek(0)

    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'asr_alignment_{task_id}.csv'
    )


# File import helper
def read_txt_file(file):
    """Read TXT file and return lines."""
    try:
        content = file.read().decode('utf-8')
        lines = [l.strip() for l in content.split('\n') if l.strip()]
        return lines
    except Exception as e:
        raise Exception(f"Error reading {file.filename}: {str(e)}")


@asr_bp.route('/import-txt', methods=['POST'])
@login_required
def import_txt_files():
    """Import TXT files for ASR and Translation."""
    if 'asr_files' not in request.files and 'trans_files' not in request.files:
        return jsonify({"error": "No files provided"}), 400
    
    asr_files = request.files.getlist('asr_files')
    trans_files = request.files.getlist('trans_files')
    
    asr_lines = []
    trans_lines = []
    
    # Process ASR files
    for f in asr_files:
        if f.filename.lower().endswith('.txt'):
            lines = read_txt_file(f)
            asr_lines.extend(lines)
    
    # Process Translation files
    for f in trans_files:
        if f.filename.lower().endswith('.txt'):
            lines = read_txt_file(f)
            trans_lines.extend(lines)
    
    return jsonify({
        "success": True,
        "asr_lines": asr_lines,
        "trans_lines": trans_lines,
        "asr_count": len(asr_lines),
        "trans_count": len(trans_lines)
    })


@asr_bp.route('/export/excel/<task_id>')
@login_required
def export_excel(task_id):
    """Export alignment results as Excel."""
    alignment_data = _stored_alignment.get("data")
    if not alignment_data:
        flash('No data to export.', 'error')
        return redirect(url_for('asr.asr_home'))

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ASR Evaluation"

        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = OpenPyXLFont(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "Sentence No.", "Test Language", "Word Count", "ASR Result (Displayed)",
            "ASR Result (Separated)", "ASR Result (No Break)", "SRR",
            "Wrong Word Count", "Sentence Score", "AI Score", "AI Reason",
            "Differences", "Wrong Words", "Translation"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        column_widths = [6, 33, 11, 33, 33, 33, 6, 11, 11, 11, 25, 33, 33, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for row_idx, d in enumerate(alignment_data, 2):
            plain_text_parts = [t for _, t in d.get("diffs", [])]
            plain_text_value = " ".join(plain_text_parts)

            row_data = [
                d.get("id", ""), d.get("true", ""), d.get("count", ""),
                d.get("asr_displayed", ""), d.get("asr_separated", ""),
                d.get("asr_nobreak", ""), d.get("srr", ""),
                d.get("wrong_count", ""), d.get("score", ""),
                d.get("ai_score", ""), d.get("ai_reason", ""),
                plain_text_value, d.get("wrong_list", ""), d.get("translation", "")
            ]
            ws.append(row_data)

            for cell in ws[row_idx]:
                cell.alignment = cell_alignment
                cell.border = thin_border

            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

        ws.freeze_panes = "A2"

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ASR_Evaluation_{timestamp}.xlsx"
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "exports", filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('asr.asr_home'))
