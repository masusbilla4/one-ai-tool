"""
Output Manager for Data Extractor.
Handles CSV and Excel export of extracted data.
"""
import csv
from typing import List, Dict, Any
from io import StringIO, BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class OutputManager:
    """Manages output of extracted data."""
    
    FIELDNAMES = ['sentence', 'word_count', 'source', 'source_type', 'timestamp']
    
    @staticmethod
    def save_csv(data: List[Dict], file_path: str = None) -> str:
        """
        Save data to CSV format.
        Returns: CSV content as string (if file_path is None) or True if saved
        """
        csv_buffer = StringIO()
        writer = csv.DictWriter(csv_buffer, fieldnames=OutputManager.FIELDNAMES, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
        
        if file_path:
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                f.write(csv_buffer.getvalue())
            return True
        
        return csv_buffer.getvalue()
    
    @staticmethod
    def save_excel(data: List[Dict], file_path: str) -> bool:
        """
        Save data to Excel format.
        Returns: True if successful
        """
        if not OPENPYXL_AVAILABLE:
            return False
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            # Header styling
            header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(OutputManager.FIELDNAMES, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row_idx, item in enumerate(data, 2):
                for col, field in enumerate(OutputManager.FIELDNAMES, 1):
                    ws.cell(row=row_idx, column=col, value=item.get(field, ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            wb.save(file_path)
            return True
        except Exception:
            return False
    
    @staticmethod
    def to_excel_bytes(data: List[Dict]) -> BytesIO:
        """
        Convert data to Excel file bytes for download.
        Returns: BytesIO object with Excel content
        """
        if not OPENPYXL_AVAILABLE:
            return None
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            # Header styling
            header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(OutputManager.FIELDNAMES, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row_idx, item in enumerate(data, 2):
                for col, field in enumerate(OutputManager.FIELDNAMES, 1):
                    ws.cell(row=row_idx, column=col, value=item.get(field, ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        except Exception:
            return None
